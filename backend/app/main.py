"""
FFP Data Validator API
Author: Fayez Ahmed
"""
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import pandas as pd
import io
import os
import uvicorn
import asyncio
from contextlib import asynccontextmanager
import openpyxl
from openpyxl.styles import PatternFill, Font
import json
from datetime import datetime
from sqlalchemy.orm import Session
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.requests import Request
from .database import engine, Base, get_db, SessionLocal
from .models import User, SystemConfig, RemoteInstance, SummaryStats, ValidRecord, InvalidRecord, UploadedFile, Division, District, Upazila, Permission, RolePermission, AuditLog, ApiUsageLog, UploadBatch
from .auth import (
    get_current_user,
    create_access_token,
    verify_password,
    get_password_hash,
    ACCESS_TOKEN_EXPIRE_MINUTES,
    hash_password,
    get_api_key
)
from .rbac import PermissionChecker
from .audit import log_audit, log_api_usage
from sqlalchemy import or_, func
from sqlalchemy.dialects.postgresql import insert
from .validator import process_dataframe
from .pdf_generator import generate_pdf_report
from .bd_geo import fuzzy_match_location, get_division_for_district
from . import auth_routes
from . import admin_routes
import urllib.parse
# Trigger hot-reload test
import zipfile
import tempfile

async def lifespan(app: FastAPI):
    os.makedirs("downloads", exist_ok=True)
    os.makedirs("uploads", exist_ok=True)
    # Enable extensions and create tables
    db_init = SessionLocal()
    try:
        from sqlalchemy import text
        db_init.execute(text("CREATE EXTENSION IF NOT EXISTS pg_trgm"))
        db_init.commit()
    except Exception as e:
        print(f"Error creating pg_trgm extension (might not be superuser or already exists): {e}")
        db_init.rollback()
    finally:
        db_init.close()

    Base.metadata.create_all(bind=engine)
    
    # Run safe column migration (add last_upload_* columns if they don't exist)
    # Run migrations and seeding
    db = SessionLocal()
    try:
        migrate_schema(db)
        migrate_json_to_db(db)
        # Seed default admin if no users exist
        if db.query(User).count() == 0:
            admin_user = User(
                username="admin",
                hashed_password=hash_password("admin123"),
                role="admin"
            )
            db.add(admin_user)
            db.commit()
            print("Default admin user created: admin / admin123")
            print("\033[91m" + "="*60)
            print("  ⚠️  SECURITY WARNING")
            print("  Default admin password 'admin123' is in use.")
            print("  CHANGE IT IMMEDIATELY via the admin panel.")
            print("="*60 + "\033[0m")
        else:
            # Warn if the hashed default password is still active
            admin = db.query(User).filter(User.username == "admin").first()
            if admin and verify_password("admin123", admin.hashed_password):
                if db.query(ValidRecord).count() > 0:
                    app.state.security_lockout = True
                    print("\033[91m" + "="*60)
                    print("  ⚠️  CRITICAL SECURITY LOCKOUT")
                    print("  Default admin password 'admin123' is active AND data exists.")
                    print("  Upload API endpoints are disabled until you change the password.")
                    print("="*60 + "\033[0m")
                else:
                    print("\033[93m" + "="*60)
                    print("  ⚠️  SECURITY WARNING")
                    print("  Default admin password 'admin123' is still active!")
                    print("  CHANGE IT IMMEDIATELY via the admin panel.")
                    print("="*60 + "\033[0m")
            
        seed_geo_data_if_empty(db)
        seed_permissions_if_empty(db) # Added this line
        
        # Load the tree into memory for the bd_geo fuzzy matching
        from . import bd_geo
        bd_geo.load_geo_data_from_db(db)
    finally:
        db.close()
    yield

def migrate_schema(db: Session):
    """Safely add any missing columns to summary_stats and users tables."""
    from sqlalchemy import text
    new_summary_columns = [
        ("last_upload_total", "INTEGER DEFAULT 0"),
        ("last_upload_valid", "INTEGER DEFAULT 0"),
        ("last_upload_invalid", "INTEGER DEFAULT 0"),
        ("last_upload_new", "INTEGER DEFAULT 0"),
        ("last_upload_updated", "INTEGER DEFAULT 0"),
        ("last_upload_duplicate", "INTEGER DEFAULT 0"),
        ("excel_valid_url", "VARCHAR"),
        ("excel_invalid_url", "VARCHAR"),
        ("pdf_invalid_url", "VARCHAR"),
    ]
    for col_name, col_def in new_summary_columns:
        try:
            db.execute(text(
                f"ALTER TABLE summary_stats ADD COLUMN IF NOT EXISTS {col_name} {col_def}"
            ))
            db.commit()
        except Exception:
            db.rollback()

    new_user_columns = [
        ("api_key_last_used", "TIMESTAMP"),
        ("api_rate_limit", "INTEGER DEFAULT 60"),
        ("api_total_limit", "INTEGER"),
        ("api_usage_count", "INTEGER DEFAULT 0"),
        ("api_ip_whitelist", "VARCHAR")
    ]
    for col_name, col_def in new_user_columns:
        try:
            db.execute(text(
                f"ALTER TABLE users ADD COLUMN IF NOT EXISTS {col_name} {col_def}"
            ))
            db.commit()
        except Exception:
            db.rollback()

    # New columns for valid_records
    try:
        db.execute(text("ALTER TABLE valid_records ADD COLUMN IF NOT EXISTS card_no VARCHAR"))
        db.commit()
    except Exception:
        db.rollback()

    # New columns for invalid_records
    invalid_cols = [
        ("card_no", "VARCHAR"),
        ("master_serial", "VARCHAR"),
        ("mobile", "VARCHAR")
    ]
    for col_name, col_def in invalid_cols:
        try:
            db.execute(text(f"ALTER TABLE invalid_records ADD COLUMN IF NOT EXISTS {col_name} {col_def}"))
            db.commit()
        except Exception:
            db.rollback()

    # Create all tables (ensure new indexes are applied if they weren't in lifespan)
    try:
        Base.metadata.create_all(bind=engine)
    except Exception:
        pass

    # Migration for upazilas: quota
    try:
        db.execute(text("ALTER TABLE upazilas ADD COLUMN IF NOT EXISTS quota INTEGER DEFAULT 0"))
        db.commit()
    except Exception:
        db.rollback()
            
    # Migration for valid_records: batch_id
    try:
        db.execute(text("ALTER TABLE valid_records ADD COLUMN IF NOT EXISTS batch_id INTEGER"))
        db.commit()
    except Exception:
        db.rollback()

    # Migration for invalid_records: batch_id
    try:
        db.execute(text("ALTER TABLE invalid_records ADD COLUMN IF NOT EXISTS batch_id INTEGER"))
        db.commit()
    except Exception:
        db.rollback()

def seed_geo_data_if_empty(db: Session):
    """Seed the divisions, districts, and upazilas tables from geo_data.json if empty."""
    import os
    import json
    
    geo_file_path = os.path.join(os.path.dirname(__file__), "geo_data.json")
    if not os.path.exists(geo_file_path):
        print("geo_data.json not found. Skipping seeding.")
        return

    try:
        if db.query(Division).count() == 0:
            with open(geo_file_path, 'r', encoding='utf-8') as f:
                divisions_data = json.load(f)
            
            div_count = 0
            dist_count = 0
            upz_count = 0
            
            for div in divisions_data:
                div_name = div["name"]
                db_div = Division(name=div_name, is_active=True)
                db.add(db_div)
                div_count += 1
                
                for dist in div["districts"]:
                    dist_name = dist["name"]
                    db_dist = District(division_name=div_name, name=dist_name, is_active=True)
                    db.add(db_dist)
                    dist_count += 1
                    
                    for upz_name in dist["upazilas"]:
                        db_upz = Upazila(division_name=div_name, district_name=dist_name, name=upz_name, is_active=True)
                        db.add(db_upz)
                        upz_count += 1
            
            db.commit()
            print(f"Successfully seeded {div_count} divisions, {dist_count} districts, and {upz_count} upazilas into the permanent database.")
    except Exception as e:
        db.rollback()
        print(f"Error seeding geo data: {e}")


def seed_permissions_if_empty(db: Session):
    """Seed the permissions and role_permissions tables if empty."""
    try:
        if db.query(Permission).count() == 0:
            perms = [
                ("upload_data", "Can upload and validate Excel files"),
                ("view_stats", "Can view statistics and dashboard"),
                ("view_geo", "Can view geographical hierarchy"),
                ("view_admin", "Can view administrative settings"),
                ("manage_users", "Can create/edit users"),
                ("manage_geo", "Can edit geographical data"),
            ]
            for name, desc in perms:
                db.add(Permission(name=name, description=desc))
            
            # Role Map
            role_map = {
                "admin": [p[0] for p in perms],
                "uploader": ["upload_data"],
                "viewer": ["view_stats", "view_geo"]
            }
            for role, pnames in role_map.items():
                for pname in pnames:
                    db.add(RolePermission(role=role, permission_name=pname))
            
            db.commit()
            print("Successfully seeded permissions and role mappings.")
    except Exception as e:
        db.rollback()
        print(f"Error seeding permissions: {e}")


def migrate_json_to_db(db: Session):
    stats_file = os.path.join("downloads", "validation_stats.json")
    if os.path.exists(stats_file):
        try:
            with open(stats_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                entries = data.get("entries", {})
                for key, val in entries.items():
                    # Check if already exists in DB
                    existing = db.query(SummaryStats).filter(
                        SummaryStats.district == val["district"],
                        SummaryStats.upazila == val["upazila"]
                    ).first()
                    if not existing:
                        # Convert ISO strings to datetime objects
                        created_at = datetime.fromisoformat(val["created_at"].rstrip("Z"))
                        updated_at = datetime.fromisoformat(val["updated_at"].rstrip("Z"))
                        
                        summary = SummaryStats(
                            division=val["division"],
                            district=val["district"],
                            upazila=val["upazila"],
                            total=val["total"],
                            valid=val["valid"],
                            invalid=val["invalid"],
                            version=val.get("version", 1),
                            filename=val.get("filename", ""),
                            pdf_url=val.get("pdf_url", ""),
                            excel_url=val.get("excel_url", ""),
                            excel_valid_url=val.get("excel_valid_url", ""),
                            excel_invalid_url=val.get("excel_invalid_url", ""),
                            created_at=created_at,
                            updated_at=updated_at,
                        )
                        db.add(summary)
                db.commit()
        except Exception as e:
            print(f"Migration error: {e}")

docs_url = None if os.environ.get("DISABLE_DOCS", "true").lower() == "true" else "/docs"
app = FastAPI(title="FFP Data Validator API", docs_url=docs_url, redoc_url=None, lifespan=lifespan)

@app.middleware("http")
async def audit_api_usage_middleware(request: Request, call_next):
    import time
    start_time = time.time()
    
    response = await call_next(request)
    
    process_time = (time.time() - start_time) * 1000
    
    # We log after the request is processed
    # Extract user if authenticated
    user_id = None
    username = None
    
    # Simple check for user in state or from auth header (without full re-verify for speed)
    # Better: Use the request.state if authentication middleware set it
    # For now, we'll try to get it from the request state if available
    user = getattr(request.state, "user", None)
    if user:
        user_id = getattr(request.state, "user_id", None)
        username = getattr(request.state, "username", None)
        if user_id is None:
            try:
                user_id = user.id
                username = user.username
            except Exception:
                pass
        
    db = SessionLocal()
    try:
        log_api_usage(
            db=db,
            user_id=user_id,
            username=username,
            method=request.method,
            path=request.url.path,
            status_code=response.status_code,
            ip_address=request.client.host if request.client else "unknown",
            latency_ms=process_time
        )
    finally:
        db.close()
        
    return response


# CORS — read allowed origins from environment (never wildcard in production)
_raw_origins = os.environ.get("ALLOWED_ORIGINS", "http://localhost,http://localhost:3000")
_allowed_origins = [o.strip() for o in _raw_origins.split(",") if o.strip()]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(auth_routes.router)
app.include_router(admin_routes.router)

from .auth import limiter
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

def get_db_rate_limit(request: Request):
    db: Session = next(get_db())
    config = db.query(SystemConfig).filter(SystemConfig.key == "rate_limit_value").first()
    limit = config.value if config else "60/minute"
    return limit

# ── Health endpoint (used by Docker healthcheck & nginx) ─────────────────────
import time as _time
_APP_START = _time.time()

@app.get("/health", include_in_schema=False)
async def health_check(db: Session = Depends(get_db)):
    """Lightweight liveness + DB probe. No auth required."""
    try:
        from sqlalchemy import text as _text
        db.execute(_text("SELECT 1"))
        db_status = "ok"
    except Exception as exc:
        db_status = f"error: {exc}"
    return {
        "status": "ok" if db_status == "ok" else "degraded",
        "db": db_status,
        "uptime_seconds": round(_time.time() - _APP_START),
        "version": "1.0.0",
    }

MAX_UPLOAD_SIZE = int(os.environ.get("MAX_UPLOAD_SIZE", 20 * 1024 * 1024))  # 20MB

@app.post("/validate", dependencies=[Depends(PermissionChecker("upload_data"))])
async def validate_excel(
    request: Request,
    file: UploadFile = File(...),
    dob_column: str = Form(...),
    nid_column: str = Form(...),
    header_row: int = Form(1),
    additional_columns: str = Form(""),
    sheet_name: str = Form(None),
    division: str = Form(None),
    district: str = Form(None),
    upazila: str = Form(None),
    is_correction: bool = Form(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # Changed from require_role to get_current_user as PermissionChecker handles role
):
    if getattr(request.app.state, "security_lockout", False):
        raise HTTPException(status_code=503, detail="Security lockout: system is running with the default admin password and contains data. Please change the admin password first.")

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed.")
        
    # Read file content
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=400, detail="File too large. Max 20MB allowed.")
    
    import uuid
    safe_filename = f"{uuid.uuid4().hex}_{os.path.basename(file.filename)}"
    # Save original file
    upload_path = os.path.join("uploads", safe_filename)
    with open(upload_path, "wb") as f:
        f.write(contents)
    
    # Track uploaded file
    db_file = UploadedFile(filename=safe_filename, original_name=file.filename, filepath=upload_path)
    db.add(db_file)
    db.commit()

    # Geo-match from filename or use manual overrides
    if division and district and upazila:
        geo = {
            "division": division.strip(),
            "district": district.strip(),
            "upazila": upazila.strip()
        }
    elif district and upazila:
        # Division not provided but district+upazila are
        from .bd_geo import get_division_for_district
        geo = {
            "division": get_division_for_district(district.strip()),
            "district": district.strip(),
            "upazila": upazila.strip()
        }
    else:
        geo = fuzzy_match_location(file.filename)
        # If fuzzy match failed, use Unknown — don't block the user from validating

    try:
        def read_and_process():
            if sheet_name and sheet_name.strip():
                df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name.strip(), header=header_row - 1, dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(contents), header=header_row - 1, dtype=str)
            return process_dataframe(df, dob_col=dob_column, nid_col=nid_column, header_row=header_row)
            
        processed_df, stats = await asyncio.to_thread(read_and_process)
    except ValueError as ve:
        raise HTTPException(status_code=400, detail=str(ve))
    except KeyError as ke:
        raise HTTPException(status_code=400, detail=f"Column not found in the uploaded file: {str(ke)}. Please verify your column selection.")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Internal server error. Please contact support.")
        
    # Parse additional columns
    add_cols = [c.strip() for c in additional_columns.split(",")] if additional_columns else []
        
    # Generate PDF
    original_filename_no_ext = os.path.splitext(file.filename)[0]
    
    # Create proper naming base: districtname_upazilaName
    if geo and geo.get("district") and geo.get("upazila") and geo.get("district") != "Unknown" and geo.get("upazila") != "Unknown":
        base_filename = f"{geo['district']}_{geo['upazila']}".replace(" ", "_").replace("/", "_")
    else:
        base_filename = original_filename_no_ext

    pdf_path = generate_pdf_report(processed_df, stats, additional_columns=add_cols, original_filename=base_filename, geo=geo)
    filename = os.path.basename(pdf_path)

    # Generate invalid-only PDF report (sent back to upazila with correction requests)
    pdf_invalid_path = generate_pdf_report(processed_df, stats, additional_columns=add_cols, original_filename=base_filename, geo=geo, invalid_only=True)
    pdf_invalid_filename = os.path.basename(pdf_invalid_path)
    
    # Generate Excel exports
    red_fill = PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
    
    # 1) Full tested output with coloring
    excel_filename = f"{base_filename}_tested.xlsx"
    excel_path = os.path.join("downloads", excel_filename)
    
    # 2) Valid-only clean output
    excel_valid_filename = f"{base_filename}_valid.xlsx"
    excel_valid_path = os.path.join("downloads", excel_valid_filename)
    
    # 3) Invalid-only output (NEW)
    excel_invalid_filename = f"{base_filename}_invalid.xlsx"
    excel_invalid_path = os.path.join("downloads", excel_invalid_filename)
    
    # Find column indices in original workbook
    dob_col_idx = None
    nid_col_idx = None
    
    if file.filename.endswith('.xls') and not file.filename.endswith('.xlsx'):
        # Fallback for strict .xls
        processed_df[dob_column] = processed_df[dob_column].astype(object)
        processed_df[nid_column] = processed_df[nid_column].astype(object)
        
        for idx in range(len(processed_df)):
            processed_df.at[idx, dob_column] = processed_df.at[idx, 'Cleaned_DOB']
            processed_df.at[idx, nid_column] = processed_df.at[idx, 'Cleaned_NID']
            
        cols_to_drop = ['Cleaned_DOB', 'Cleaned_NID', 'DOB_Year', 'Status', 'Message', 'Excel_Row', 'Extracted_Name', 'Card_No', 'Master_Serial', 'Mobile']
        export_df = processed_df.drop(columns=[c for c in cols_to_drop if c in processed_df.columns])
        export_df.to_excel(excel_path, index=False, engine='openpyxl')
        
        # Color the rows in the generated .xlsx
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        for idx, row in processed_df.iterrows():
            r = idx + 2
            status = row['Status']
            if status == 'error':
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = red_fill
            elif status == 'warning':
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = yellow_fill
        wb.save(excel_path)
        
        valid_mask = processed_df['Status'] != 'error'
        with pd.ExcelWriter(excel_valid_path, engine='openpyxl') as writer:
            export_df[valid_mask].to_excel(writer, index=False, sheet_name='Valid Records')
            ws = writer.sheets['Valid Records']
            nikosh_font = Font(name='Nikosh', size=11)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = nikosh_font
        
        invalid_mask = processed_df['Status'] == 'error'
        with pd.ExcelWriter(excel_invalid_path, engine='openpyxl') as writer:
            export_df[invalid_mask].to_excel(writer, index=False, sheet_name='Invalid Records')
            ws = writer.sheets['Invalid Records']
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = nikosh_font
        
    else:
        # Standard .xlsx handling
        # 1. Tested Workbook (Retain Original Formatting)
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb[sheet_name.strip()] if sheet_name and sheet_name.strip() in wb.sheetnames else wb.active
            
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if str(val).strip() == dob_column.strip():
                dob_col_idx = col_idx
            if str(val).strip() == nid_column.strip():
                nid_col_idx = col_idx
                
        for row in processed_df.itertuples(index=False):
            r = int(row.Excel_Row)
            status = row.Status
            
            # Update cells safely (skip MergedCells which are read-only)
            if dob_col_idx:
                c_dob = ws.cell(row=r, column=dob_col_idx)
                if type(c_dob).__name__ != 'MergedCell':
                    c_dob.value = row.Cleaned_DOB
                    c_dob.number_format = '@'
                    
            if nid_col_idx:
                c_nid = ws.cell(row=r, column=nid_col_idx)
                if type(c_nid).__name__ != 'MergedCell':
                    c_nid.value = row.Cleaned_NID
                    c_nid.number_format = '@'
            
            # Apply highlighting safely
            if status == 'error':
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if type(cell).__name__ != 'MergedCell':
                        cell.fill = red_fill
            elif status == 'warning':
                for c in range(1, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c)
                    if type(cell).__name__ != 'MergedCell':
                        cell.fill = yellow_fill
        
        # Apply Nikosh font to the tested workbook
        nikosh_font = Font(name='Nikosh', size=11)
        for row in ws.iter_rows():
            for cell in row:
                cell.font = nikosh_font
                        
        wb.save(excel_path)
        
        # 2. Valid and Invalid Workbooks (Fast Pandas Export instead of openpyxl row deletion)
        # Note: We swap the raw DOB/NID data with the cleaned versions for the final downloads
        export_df = processed_df.copy()
        export_df[dob_column] = export_df['Cleaned_DOB']
        export_df[nid_column] = export_df['Cleaned_NID']
        
        cols_to_drop = ['Cleaned_DOB', 'Cleaned_NID', 'DOB_Year', 'Status', 'Message', 'Excel_Row', 'Extracted_Name', 'Card_No', 'Master_Serial', 'Mobile']
        export_df = export_df.drop(columns=[c for c in cols_to_drop if c in export_df.columns])
        
        valid_mask = processed_df['Status'] != 'error'
        with pd.ExcelWriter(excel_valid_path, engine='openpyxl') as writer:
            export_df[valid_mask].to_excel(writer, index=False, sheet_name='Valid Records')
            ws = writer.sheets['Valid Records']
            nikosh_font = Font(name='Nikosh', size=11)
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = nikosh_font
        
        invalid_mask = processed_df['Status'] == 'error'
        with pd.ExcelWriter(excel_invalid_path, engine='openpyxl') as writer:
            export_df[invalid_mask].to_excel(writer, index=False, sheet_name='Invalid Records')
            ws = writer.sheets['Invalid Records']
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = nikosh_font
    
    # Prepare preview data (first 50 rows)
    preview_df = processed_df.head(50).replace({float('nan'): None})
    preview_data = preview_df.to_dict(orient="records")
    
    # Calculate valid/invalid counts
    error_count = int((processed_df['Status'] == 'error').sum())
    valid_count = stats['total_rows'] - error_count
    
    # ── Database Persistence (NID-Aware Upsert) ──
    
    # Create UploadBatch entry
    batch = UploadBatch(
        filename=file.filename,
        original_name=file.filename,
        uploader_id=current_user.id,
        username=current_user.username,
        division=geo["division"],
        district=geo["district"],
        upazila=geo["upazila"],
        total_rows=stats['total_rows'],
        valid_count=valid_count,
        invalid_count=error_count,
        status="completed"
    )
    db.add(batch)
    db.commit()
    db.refresh(batch)

    # 1. Upsert ValidRecords — NID is unique across all of Bangladesh
    valid_rows = processed_df[processed_df['Status'] != 'error']
    new_count = 0
    updated_count = 0
    cross_upazila_duplicates = []  # Track NIDs found in different upazilas
    
    # Get the current version for this district+upazila
    summary = db.query(SummaryStats).filter(
        SummaryStats.district == geo["district"],
        SummaryStats.upazila == geo["upazila"]
    ).first()
    current_version = (summary.version + 1) if summary else 1
    
    # Bulk check existing NIDs
    valid_nids = [str(row['Cleaned_NID']).strip() for _, row in valid_rows.iterrows() if pd.notna(row['Cleaned_NID']) and str(row['Cleaned_NID']).strip()]
    existing_records = []
    if valid_nids:
        # Query in chunks if too many NIDs
        for i in range(0, len(valid_nids), 5000):
            chunk = valid_nids[i:i+5000]
            existing_records.extend(db.query(ValidRecord.nid, ValidRecord.district, ValidRecord.upazila).filter(ValidRecord.nid.in_(chunk)).all())
    existing_map = {r.nid: r for r in existing_records}

    insert_data = []    
    for _, row in valid_rows.iterrows():
        nid = str(row['Cleaned_NID']).strip()
        if not nid:
            continue
            
        row_dict = row.to_dict()
        # Sanitize NaN values for JSON storage
        row_dict = {k: (None if isinstance(v, float) and pd.isna(v) else v) for k, v in row_dict.items()}
        
        name_val = row.get('Extracted_Name', 'Unknown')
        if pd.isna(name_val): name_val = 'Unknown'
        dob_val = row.get('Cleaned_DOB', '')
        if pd.isna(dob_val): dob_val = ''
        
        if nid in existing_map:
            existing = existing_map[nid]
            if existing.upazila != geo["upazila"] or existing.district != geo["district"]:
                cross_upazila_duplicates.append({
                    "nid": nid,
                    "name": name_val,
                    "previous_district": existing.district,
                    "previous_upazila": existing.upazila,
                    "new_district": geo["district"],
                    "new_upazila": geo["upazila"]
                })
            updated_count += 1
        else:
            new_count += 1
            existing_map[nid] = existing_records # prevent counting twice if duplicate inside the sheet itself
            
        insert_data.append({
            "nid": nid,
            "dob": dob_val,
            "name": name_val,
            "division": geo["division"],
            "district": geo["district"],
            "upazila": geo["upazila"],
            "card_no": row.get('Card_No', ''),
            "source_file": file.filename,
            "batch_id": batch.id,
            "upload_batch": current_version,
            "data": row_dict,
            "updated_at": datetime.utcnow()
        })
        
    if insert_data:
        # Chunked bulk upsert
        for i in range(0, len(insert_data), 2000):
            chunk = insert_data[i:i+2000]
            stmt = insert(ValidRecord).values(chunk)
            stmt = stmt.on_conflict_do_update(
                index_elements=['nid'],
                set_={
                    'dob': stmt.excluded.dob,
                    'name': stmt.excluded.name,
                    'division': stmt.excluded.division,
                    'district': stmt.excluded.district,
                    'upazila': stmt.excluded.upazila,
                    'card_no': stmt.excluded.card_no,
                    'source_file': stmt.excluded.source_file,
                    'batch_id': stmt.excluded.batch_id,
                    'upload_batch': stmt.excluded.upload_batch,
                    'data': stmt.excluded.data,
                    'updated_at': stmt.excluded.updated_at
                }
            )
            # Execute async so other users aren't blocked completely out of db connections
            db.execute(stmt)
    

    # --- InvalidRecord inserting ---
    invalid_rows = processed_df[processed_df['Status'] == 'error']
    invalid_insert_data = []
    for _, row in invalid_rows.iterrows():
        nid_val = str(row.get('Cleaned_NID', '')).strip()
        if pd.isna(nid_val): nid_val = ''
        
        row_dict = row.to_dict()
        row_dict = {k: (None if isinstance(v, float) and pd.isna(v) else v) for k, v in row_dict.items()}
        
        name_val = row.get('Extracted_Name', 'Unknown')
        if pd.isna(name_val): name_val = 'Unknown'
        dob_val = row.get('Cleaned_DOB', '')
        if pd.isna(dob_val): dob_val = ''
        
        invalid_insert_data.append({
            "nid": nid_val,
            "dob": dob_val,
            "name": name_val,
            "division": geo["division"],
            "district": geo["district"],
            "upazila": geo["upazila"],
            "card_no": row.get('Card_No', ''),
            "master_serial": row.get('Master_Serial', ''),
            "mobile": row.get('Mobile', ''),
            "source_file": file.filename,
            "batch_id": batch.id,
            "upload_batch": current_version,
            "error_message": str(row.get('Message', 'Unknown Error')),
            "data": row_dict
        })
        
    if invalid_insert_data:
        for i in range(0, len(invalid_insert_data), 2000):
            chunk = invalid_insert_data[i:i+2000]
            db.bulk_insert_mappings(InvalidRecord, chunk)
    # --- End InvalidRecord inserting ---
    
    # --- Iterative Correction Logic: Clear old errors ---
    if is_correction:
        # Multi-layered matching to delete corrected records from InvalidRecord
        # Match by NID, Card No, Name, or Mobile — any identifier match counts
        
        valid_nids = [str(r['nid']) for r in insert_data if r.get('nid')]
        valid_cards = [str(r['card_no']) for r in insert_data if r.get('card_no')]
        valid_names = [str(r['name']) for r in insert_data if r.get('name') and r['name'] != 'Unknown']
        valid_mobiles = []
        for r in insert_data:
            mob = r.get('data', {}).get('Mobile', '') if isinstance(r.get('data'), dict) else ''
            if mob and str(mob).strip():
                valid_mobiles.append(str(mob).strip())
        
        match_conditions = []
        from sqlalchemy import or_
        if valid_nids:
            match_conditions.append(InvalidRecord.nid.in_(valid_nids))
        if valid_cards:
            match_conditions.append(InvalidRecord.card_no.in_(valid_cards))
        if valid_names:
            match_conditions.append(InvalidRecord.name.in_(valid_names))
        if valid_mobiles:
            match_conditions.append(InvalidRecord.mobile.in_(valid_mobiles))
        
        if match_conditions:
            db.query(InvalidRecord).filter(
                InvalidRecord.upazila == geo["upazila"],
                InvalidRecord.district == geo["district"],
                or_(*match_conditions)
            ).delete(synchronize_session=False)
            db.commit()
    # --- End Iterative Correction Logic ---
    
    # Update batch with results
    batch.new_records = new_count

    batch.updated_records = updated_count
    db.commit()

    # 2. Re-calculate SummaryStats from scratch (absolute truth)
    total_valid = db.query(ValidRecord).filter(
        ValidRecord.upazila == geo["upazila"],
        ValidRecord.district == geo["district"]
    ).count()
    
    total_invalid = db.query(InvalidRecord).filter(
        InvalidRecord.upazila == geo["upazila"],
        InvalidRecord.district == geo["district"]
    ).count()

    if summary:
        summary.valid = total_valid
        summary.invalid = total_invalid
        # total represents the total "slots" or beneficiaries we are tracking
        # This is essentially valid + invalid (errors are still meant to be valid records)
        summary.total = total_valid + total_invalid
        
        summary.last_upload_total = stats['total_rows']
        summary.last_upload_valid = valid_count
        summary.last_upload_invalid = error_count
        summary.last_upload_new = new_count
        summary.last_upload_updated = updated_count
        summary.last_upload_duplicate = len(cross_upazila_duplicates)
        summary.version = current_version
        summary.filename = file.filename
        summary.pdf_url = f"/api/downloads/{urllib.parse.quote(filename)}"
        summary.pdf_invalid_url = f"/api/downloads/{urllib.parse.quote(pdf_invalid_filename)}" if pdf_invalid_filename else ""
        summary.excel_url = f"/api/downloads/{urllib.parse.quote(excel_filename)}" if excel_filename else ""
        summary.excel_valid_url = f"/api/downloads/{urllib.parse.quote(excel_valid_filename)}" if excel_valid_filename else ""
        summary.excel_invalid_url = f"/api/downloads/{urllib.parse.quote(excel_invalid_filename)}" if excel_invalid_filename else ""
    else:
        summary = SummaryStats(
            division=geo["division"],
            district=geo["district"],
            upazila=geo["upazila"],
            total=total_valid + total_invalid,
            valid=total_valid,
            invalid=total_invalid,
            last_upload_total=stats['total_rows'],
            last_upload_valid=valid_count,
            last_upload_invalid=error_count,
            last_upload_new=new_count,
            last_upload_updated=updated_count,
            last_upload_duplicate=len(cross_upazila_duplicates),
            filename=file.filename,
            pdf_url=f"/api/downloads/{urllib.parse.quote(filename)}",
            pdf_invalid_url=f"/api/downloads/{urllib.parse.quote(pdf_invalid_filename)}" if pdf_invalid_filename else "",
            excel_url=f"/api/downloads/{urllib.parse.quote(excel_filename)}" if excel_filename else "",
            excel_valid_url=f"/api/downloads/{urllib.parse.quote(excel_valid_filename)}" if excel_valid_filename else "",
            excel_invalid_url=f"/api/downloads/{urllib.parse.quote(excel_invalid_filename)}" if excel_invalid_filename else "",
        )
        db.add(summary)
    
    db.commit()
    db.refresh(summary)
    
    log_audit(
        db, 
        current_user, 
        "CREATE", 
        "upload_batch", 
        batch.id, 
        new_data={
            "action": "file_upload_validation", 
            "filename": file.filename,
            "total_rows": stats['total_rows'],
            "valid": valid_count,
            "invalid": error_count,
            "new": new_count,
            "updated": updated_count,
            "geo": geo,
            "summary_id": summary.id
        }
    )
    
    return {
        "summary": stats,
        "geo": geo,
        "valid_count": valid_count,
        "invalid_count": error_count,
        "new_records": new_count,
        "updated_records": updated_count,
        "cross_upazila_duplicates": cross_upazila_duplicates,
        "batch_id": batch.id,
        "version": summary.version,
        "updated_at": summary.updated_at.isoformat() + "Z",
        "pdf_url": f"/api/downloads/{urllib.parse.quote(filename)}",
        "pdf_invalid_url": f"/api/downloads/{urllib.parse.quote(pdf_invalid_filename)}",
        "excel_url": f"/api/downloads/{urllib.parse.quote(excel_filename)}",
        "excel_valid_url": f"/api/downloads/{urllib.parse.quote(excel_valid_filename)}",
        "excel_invalid_url": f"/api/downloads/{urllib.parse.quote(excel_invalid_filename)}",
        "preview_data": preview_data
    }

@app.post("/preview", dependencies=[Depends(PermissionChecker("upload_data"))])
async def preview_validation(
    file: UploadFile = File(...),
    dob_column: str = Form(...),
    nid_column: str = Form(...),
    header_row: int = Form(1),
    sheet_name: str = Form(None),
):
    """Dry-run validation of the first 10 rows to catch column mismatches early.
    Returns blocked=true if >50% of preview rows are invalid."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are allowed.")
        
    contents = await file.read()
    try:
        def read_and_process():
            if sheet_name and sheet_name.strip():
                df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name.strip(), header=header_row - 1, nrows=10, dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(contents), header=header_row - 1, nrows=10, dtype=str)
            return process_dataframe(df, dob_col=dob_column, nid_col=nid_column, header_row=header_row)
            
        processed_df, stats = await asyncio.to_thread(read_and_process)
        preview_data = processed_df.replace({float('nan'): None}).to_dict(orient="records")
        
        # Calculate invalid percentage and blocking threshold
        total = len(preview_data)
        invalid = sum(1 for r in preview_data if r.get('Status') == 'error')
        invalid_pct = round((invalid / total) * 100, 1) if total > 0 else 0
        blocked = invalid_pct > 50
        
        return {
            "preview": preview_data, 
            "summary": stats,
            "invalid_pct": invalid_pct,
            "blocked": blocked
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/statistics/history", dependencies=[Depends(PermissionChecker("view_stats"))])
async def get_upazila_history(
    district: str,
    upazila: str,
    db: Session = Depends(get_db)
):
    """Return upload history for a specific upazila, validated against official names."""
    # First, let's make sure we are using the official names if possible
    # This handles the case where the frontend sends a slightly mismatched name
    official = db.query(Upazila).filter(
        (Upazila.name.ilike(upazila)) & (Upazila.district_name.ilike(district))
    ).first()
    
    target_upazila = official.name if official else upazila
    target_district = official.district_name if official else district

    batches = db.query(UploadBatch).filter(
        UploadBatch.district == target_district,
        UploadBatch.upazila == target_upazila
    ).order_by(UploadBatch.created_at.desc()).all()
    return batches

@app.delete("/batches/{batch_id}", dependencies=[Depends(PermissionChecker("manage_geo"))])
async def delete_batch(
    batch_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete an upload batch and its associated valid records. Updates summary stats."""
    batch = db.query(UploadBatch).filter(UploadBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
        
    if batch.status == "deleted":
        return {"message": "Batch already deleted"}

    # 1. Update SummaryStats
    summary = db.query(SummaryStats).filter(
        SummaryStats.district == batch.district,
        SummaryStats.upazila == batch.upazila
    ).first()
    
    if summary:
        # We decrease the counts. 
        # CAUTION: If records were updated by a later batch, this hard delete might be tricky.
        # For now, we delete all records matching this batch.
        
        # Count records to be deleted
        valid_to_delete = db.query(ValidRecord).filter(ValidRecord.batch_id == batch_id).count()
        
        summary.valid = max(0, summary.valid - batch.new_records)
        summary.invalid = max(0, summary.invalid - batch.invalid_count)
        summary.total = summary.valid + summary.invalid
        summary.version += 1
        
    # 2. Delete the valid records
    db.query(ValidRecord).filter(ValidRecord.batch_id == batch_id).delete()
    db.query(InvalidRecord).filter(InvalidRecord.batch_id == batch_id).delete()
    
    # 3. Mark batch as deleted instead of hard delete for audit
    batch.status = "deleted"
    db.commit()
    
    log_audit(db, current_user, "DELETE", "upload_batch", batch_id, old_data={"filename": batch.filename})
    
    return {"message": "Batch deleted successfully", "deleted_valid": batch.new_records}

# ─────────────────────────────────────────────────────────────────────────────
# LIVE EXPORT  — always fresh from DB, all versions merged

# ─────────────────────────────────────────────────────────────────────────────
# LIVE EXPORT HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _get_live_records_df(db: Session, division: str, district: str, upazila: str, is_invalid: bool = False):
    """Fetch live records and return a formatted DataFrame."""
    if is_invalid:
        records = (
            db.query(InvalidRecord)
            .filter(
                InvalidRecord.division == division,
                InvalidRecord.district == district,
                InvalidRecord.upazila  == upazila,
            )
            .order_by(InvalidRecord.id.desc())
            .all()
        )
    else:
        records = (
            db.query(ValidRecord)
            .filter(
                ValidRecord.division == division,
                ValidRecord.district == district,
                ValidRecord.upazila  == upazila,
            )
            .order_by(ValidRecord.nid)
            .all()
        )

    if not records:
        return None

    rows = []
    for r in records:
        # Base columns
        row = {
            "Excel_Row": r.data.get("Excel_Row", "") if r.data else "",
            "NID":      r.nid,
            "Cleaned_NID": r.nid,
            "DOB":      r.dob,
            "Cleaned_DOB": r.dob,
            "Name":     r.name,
            "Division": r.division,
            "District": r.district,
            "Upazila":  r.upazila,
            "Batch_ID": r.batch_id,
            "Source_File": r.source_file,
        }
        
        if is_invalid:
            row["Status"] = "error"
            row["Message"] = r.error_message
        else:
            row["Status"] = "valid"
            row["Message"] = ""

        # Overlay original columns from JSON data
        if r.data and isinstance(r.data, dict):
            for k, v in r.data.items():
                if k not in row and k not in ["Status", "Message", "Excel_Row", "Cleaned_DOB", "Cleaned_NID"]:
                    row[k] = v
        rows.append(row)

    return pd.DataFrame(rows)

def _save_live_excel_nikosh(df: pd.DataFrame, path: str, sheet_name: str, is_valid: bool = True):
    """Save DataFrame to Excel with Nikosh font and column filtering."""
    if is_valid:
        # Columns to remove for valid export as requested
        exclude = ["Excel_Row", "NID", "Cleaned_NID", "DOB", "Cleaned_DOB", 
                   "Name", "Status", "Message", "Division", "District", 
                   "Upazila", "Batch_ID", "Source_File"]
    else:
        # Columns to remove for invalid export
        exclude = ["Cleaned_DOB", "Cleaned_NID", "Status"]
        
    export_df = df.drop(columns=[c for c in exclude if c in df.columns])
    
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        
        from openpyxl.styles import Font
        nikosh_font = Font(name='Nikosh', size=11)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = nikosh_font

@app.get("/upazila/live-export-invalid", dependencies=[Depends(PermissionChecker("view_stats"))])
async def upazila_live_export_invalid(
    division: str,
    district: str,
    upazila: str,
    fmt: str = "pdf",   # xlsx | pdf
    db: Session = Depends(get_db),
):
    """Stream a freshly generated export of ALL invalid records for the upazila."""
    df = _get_live_records_df(db, division, district, upazila, is_invalid=True)
    if df is None:
        raise HTTPException(status_code=404, detail="No invalid records found for this upazila")

    safe_name = f"{district}_{upazila}".replace(" ", "_").replace("/", "_")
    os.makedirs("downloads/live", exist_ok=True)

    if fmt == "pdf":
        stats = {"total_rows": len(df), "issues": len(df), "converted_nid": 0}
        geo   = {"division": division, "district": district, "upazila": upazila}
        path = generate_pdf_report(
            df, stats,
            additional_columns=[c for c in df.columns if c not in
                                 ["Status","Message","Excel_Row","Cleaned_DOB","Cleaned_NID", "NID", "DOB", "Batch_ID", "Source_File", "Division", "District", "Upazila"]],
            output_dir="downloads/live",
            original_filename=safe_name + "_live_invalid",
            geo=geo,
            invalid_only=True
        )
        media = "application/pdf"
        dl_name = f"{safe_name}_live_invalid.pdf"
    else:
        path = os.path.join("downloads", "live", f"{safe_name}_live_invalid.xlsx")
        _save_live_excel_nikosh(df, path, "Invalid Records", is_valid=False)
        media   = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        dl_name = f"{safe_name}_live_invalid.xlsx"

    from fastapi.responses import FileResponse
    return FileResponse(path, media_type=media, filename=dl_name)

@app.get("/upazila/live-export", dependencies=[Depends(PermissionChecker("view_stats"))])
async def upazila_live_export(
    division: str,
    district: str,
    upazila: str,
    fmt: str = "xlsx",   # xlsx | pdf
    db: Session = Depends(get_db),
):
    """Stream a freshly generated export of ALL valid records for the upazila."""
    df = _get_live_records_df(db, division, district, upazila, is_invalid=False)
    if df is None:
        raise HTTPException(status_code=404, detail="No valid records found for this upazila")

    safe_name = f"{district}_{upazila}".replace(" ", "_").replace("/", "_")
    os.makedirs("downloads/live", exist_ok=True)

    if fmt == "pdf":
        stats = {"total_rows": len(df), "issues": 0, "converted_nid": 0}
        geo   = {"division": division, "district": district, "upazila": upazila}
        # Add metadata for PDF colouring
        df["Status"]  = "success"
        df["Message"] = "Valid record"
        df["Excel_Row"] = range(2, len(df) + 2)
        df["Cleaned_DOB"] = df["DOB"]
        df["Cleaned_NID"] = df["NID"]
        path = generate_pdf_report(
            df, stats,
            additional_columns=[c for c in df.columns if c not in
                                 ["Status","Message","Excel_Row","Cleaned_DOB","Cleaned_NID"]],
            output_dir="downloads/live",
            original_filename=safe_name + "_live_valid",
            geo=geo,
        )
        media = "application/pdf"
        dl_name = f"{safe_name}_live_valid.pdf"
    else:
        path = os.path.join("downloads", "live", f"{safe_name}_live_valid.xlsx")
        _save_live_excel_nikosh(df, path, "Valid Records", is_valid=True)
        media   = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        dl_name = f"{safe_name}_live_valid.xlsx"

    return FileResponse(path, media_type=media, filename=dl_name)


# ─────────────────────────────────────────────────────────────────────────────
# RE-CHECK  — run fraud detection on already-stored valid records
# ─────────────────────────────────────────────────────────────────────────────

@app.post("/upazila/recheck", dependencies=[Depends(PermissionChecker("view_admin"))])
async def upazila_recheck(
    division: str,
    district: str,
    upazila: str,
    fmt: str = "xlsx",   # xlsx | pdf
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Re-run NID fraud checks on every stored valid record for the upazila.
    Returns a report of suspicious records.  Admin-only.
    """
    from .validator import check_fake_nid

    records = (
        db.query(ValidRecord)
        .filter(
            ValidRecord.division == division,
            ValidRecord.district == district,
            ValidRecord.upazila  == upazila,
        )
        .all()
    )

    if not records:
        raise HTTPException(status_code=404, detail="No records found for this upazila")

    flagged = []
    for r in records:
        is_fake, reason = check_fake_nid(r.nid or "")
        if is_fake:
            row = {
                "NID":      r.nid,
                "DOB":      r.dob,
                "Name":     r.name,
                "Batch_ID": r.batch_id,
                "Source_File": r.source_file,
                "Fraud_Reason": reason,
            }
            if r.data and isinstance(r.data, dict):
                for k, v in r.data.items():
                    if k not in row:
                        row[k] = v
            flagged.append(row)

    log_audit(db, current_user, "RECHECK", "valid_records", None,
              new_data={"upazila": upazila, "total": len(records), "flagged": len(flagged)})

    if fmt == "json":
        return {"total_checked": len(records), "flagged_count": len(flagged), "flagged": flagged}

    safe_name = f"{district}_{upazila}".replace(" ", "_").replace("/", "_")
    os.makedirs("downloads/recheck", exist_ok=True)

    if not flagged:
        # Return a 200 JSON when nothing flagged — no file needed
        return {"total_checked": len(records), "flagged_count": 0,
                "message": "No suspicious NIDs found in stored records"}

    df = pd.DataFrame(flagged)

    if fmt == "pdf":
        df["Status"]    = "error"
        df["Message"]   = df["Fraud_Reason"]
        df["Excel_Row"] = range(2, len(df) + 2)
        df["Cleaned_DOB"] = df["DOB"]
        df["Cleaned_NID"] = df["NID"]
        stats = {"total_rows": len(records), "issues": len(flagged), "converted_nid": 0}
        geo   = {"division": division, "district": district, "upazila": upazila}
        path  = generate_pdf_report(
            df, stats,
            additional_columns=["Fraud_Reason"],
            output_dir="downloads/recheck",
            original_filename=safe_name + "_recheck",
            geo=geo,
            invalid_only=False,
        )
        media   = "application/pdf"
        dl_name = f"{safe_name}_fraud_report.pdf"
    else:
        path = os.path.join("downloads", "recheck", f"{safe_name}_recheck.xlsx")
        df.to_excel(path, index=False)
        media   = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        dl_name = f"{safe_name}_fraud_report.xlsx"

    return FileResponse(path, media_type=media, filename=dl_name)


# ─────────────────────────────────────────────────────────────────────────────
# BATCH FILES  — versioned archive links per upazila
# ─────────────────────────────────────────────────────────────────────────────

@app.get("/upazila/batch-files", dependencies=[Depends(PermissionChecker("view_stats"))])
async def upazila_batch_files(
    division: str,
    district: str,
    upazila: str,
    db: Session = Depends(get_db),
):
    """Return all upload batches for this upazila with their archived file URLs."""
    batches = (
        db.query(UploadBatch)
        .filter(
            UploadBatch.division == division,
            UploadBatch.district == district,
            UploadBatch.upazila  == upazila,
            UploadBatch.status   == "completed",
        )
        .order_by(UploadBatch.created_at.desc())
        .all()
    )

    result = []
    for b in batches:
        # Reconstruct the archived filenames using the same naming logic as the upload endpoint
        safe = f"{b.district}_{b.upazila}".replace(" ", "_").replace("/", "_")
        entry = {
            "batch_id":      b.id,
            "filename":      b.original_name or b.filename,
            "uploaded_at":   b.created_at.isoformat() + "Z",
            "uploaded_by":   b.username,
            "total_rows":    b.total_rows,
            "valid_count":   b.valid_count,
            "invalid_count": b.invalid_count,
            "new_records":   b.new_records,
            "updated_records": b.updated_records,
            # Archive URLs — these are the static files saved at upload time
            "valid_url":   f"/api/downloads/{urllib.parse.quote(safe + '_valid.xlsx')}",
            "invalid_url": f"/api/downloads/{urllib.parse.quote(safe + '_invalid.xlsx')}",
            "pdf_url":     f"/api/downloads/{urllib.parse.quote(safe + '_validation_Report.pdf')}",
            "pdf_invalid_url": f"/api/downloads/{urllib.parse.quote(safe + '_invalid_Report.pdf')}",
        }
        # Check which files actually exist on disk
        for key in ("valid_url", "invalid_url", "pdf_url", "pdf_invalid_url"):
            local = os.path.join("downloads", urllib.parse.unquote(os.path.basename(entry[key])))
            entry[key] = entry[key] if os.path.exists(local) else None
        result.append(entry)

    return {"batches": result, "total": len(result)}


@app.get("/downloads/valid-zip", dependencies=[Depends(PermissionChecker("view_stats"))])
async def download_all_valid_zip(db: Session = Depends(get_db)):
    """Zip all live valid records for all upazilas, applying Nikosh font and filtering."""
    # Query SummaryStats to know which upazilas have valid records
    entries = db.query(SummaryStats).filter(
        SummaryStats.valid > 0
    ).order_by(SummaryStats.division, SummaryStats.district, SummaryStats.upazila).all()

    if not entries:
        raise HTTPException(status_code=404, detail="No valid records found in system")

    os.makedirs("downloads/temp_bulk", exist_ok=True)
    zip_filename = f"all_live_valid_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    zip_path = os.path.join("downloads", zip_filename)
    
    print(f"DEBUG: Starting valid-zip generation for {len(entries)} entries. Path: {zip_path}")
    try:
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for entry in entries:
                df = _get_live_records_df(db, entry.division, entry.district, entry.upazila, is_invalid=False)
                if df is None:
                    continue

                div  = str(entry.division or "Unknown").replace(" ", "_").replace("/", "_")
                dist = str(entry.district or "Unknown").replace(" ", "_").replace("/", "_")
                upz  = str(entry.upazila  or "Unknown").replace(" ", "_").replace("/", "_")

                # Temp file for this upazila
                temp_file = os.path.join("downloads/temp_bulk", f"{dist}_{upz}_valid.xlsx")
                _save_live_excel_nikosh(df, temp_file, "Valid Records", is_valid=True)

                # Add to zip: Division/DistrictName_UpazilaName_valid.xlsx
                zipf.write(temp_file, arcname=f"{div}/{dist}_{upz}_valid.xlsx")
                
                # Cleanup individual file after adding to zip
                if os.path.exists(temp_file):
                    os.remove(temp_file)

        if not os.path.exists(zip_path) or os.path.getsize(zip_path) < 100: # zip header is ~22 bytes
            if os.path.exists(zip_path): os.remove(zip_path)
            raise HTTPException(status_code=404, detail="Zip file generation failed or empty")

        return FileResponse(zip_path, media_type="application/zip", filename="All_Live_Valid_Records.zip")
    except Exception as e:
        if os.path.exists(zip_path): os.remove(zip_path)
        raise HTTPException(status_code=500, detail=f"Failed to create live valid zip: {str(e)}")
    finally:
        # Final cleanup of temp dir if needed
        import shutil
        if os.path.exists("downloads/temp_bulk"):
            shutil.rmtree("downloads/temp_bulk")


@app.get("/downloads/invalid-zip", dependencies=[Depends(PermissionChecker("view_stats"))])
async def download_all_invalid_zip(db: Session = Depends(get_db)):
    """Zip all live invalid records for all upazilas, applying Nikosh font."""
    entries = db.query(SummaryStats).filter(
        SummaryStats.invalid > 0
    ).order_by(SummaryStats.division, SummaryStats.district, SummaryStats.upazila).all()

    if not entries:
        raise HTTPException(status_code=404, detail="No invalid records found in system")

    os.makedirs("downloads/temp_bulk_invalid", exist_ok=True)
    zip_filename = f"all_live_invalid_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    zip_path = os.path.join("downloads", zip_filename)

    print(f"DEBUG: Starting invalid-zip generation for {len(entries)} entries. Path: {zip_path}")
    try:
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as zipf:
            for entry in entries:
                df = _get_live_records_df(db, entry.division, entry.district, entry.upazila, is_invalid=True)
                if df is None:
                    continue

                div  = str(entry.division or "Unknown").replace(" ", "_").replace("/", "_")
                dist = str(entry.district or "Unknown").replace(" ", "_").replace("/", "_")
                upz  = str(entry.upazila  or "Unknown").replace(" ", "_").replace("/", "_")

                temp_file = os.path.join("downloads/temp_bulk_invalid", f"{dist}_{upz}_invalid.xlsx")
                _save_live_excel_nikosh(df, temp_file, "Invalid Records", is_valid=False)

                zipf.write(temp_file, arcname=f"{div}/{dist}_{upz}_invalid.xlsx")
                
                if os.path.exists(temp_file):
                    os.remove(temp_file)

        if not os.path.exists(zip_path) or os.path.getsize(zip_path) < 100:
            if os.path.exists(zip_path): os.remove(zip_path)
            raise HTTPException(status_code=404, detail="Zip file generation failed or empty")

        return FileResponse(zip_path, media_type="application/zip", filename="All_Live_Invalid_Records.zip")
    except Exception as e:
        if os.path.exists(zip_path): os.remove(zip_path)
        raise HTTPException(status_code=500, detail=f"Failed to create live invalid zip: {str(e)}")
    finally:
        import shutil
        if os.path.exists("downloads/temp_bulk_invalid"):
            shutil.rmtree("downloads/temp_bulk_invalid")

@app.get("/downloads/{filename}")
async def download_file(filename: str, request: Request):
    # │ Security: strip any path separators to prevent directory traversal
    safe_name = os.path.basename(filename)
    print(f"DEBUG: Download request for file: {safe_name}")
    file_path  = os.path.join("downloads", safe_name)
    # Double-check the resolved path is still inside 'downloads/'
    downloads_dir = os.path.abspath("downloads")
    abs_path      = os.path.abspath(file_path)
    if not abs_path.startswith(downloads_dir + os.sep):
        raise HTTPException(status_code=400, detail="Invalid file path")
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="File not found")

    if safe_name.endswith('.pdf'):
        media_type = "application/pdf"
    elif safe_name.endswith('.zip'):
        media_type = "application/zip"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(abs_path, media_type=media_type, filename=safe_name)

@app.get("/statistics", dependencies=[Depends(PermissionChecker("view_stats"))])
async def get_statistics(
    request: Request,
    has_invalid: bool = False,
    division: str = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Return accumulated stats derived from official Upazila table + SummaryStats.
    
    Optional query params:
    - has_invalid: If true, only return upazilas that have invalid > 0
    - division: Filter by division name
    """
    import hashlib

    # 1. Fetch active upazilas and join with their summary stats
    query = db.query(
        Upazila,
        SummaryStats
    ).outerjoin(
        SummaryStats, 
        (Upazila.name == SummaryStats.upazila) & (Upazila.district_name == SummaryStats.district)
    ).filter(
        Upazila.is_active == True
    )

    # Apply server-side filters
    if division:
        query = query.filter(Upazila.division_name == division)
    if has_invalid:
        query = query.filter(SummaryStats.invalid > 0)

    entries_query = query.order_by(
        Upazila.division_name, Upazila.district_name, Upazila.name
    ).all()

    # Convert to objects for easier processing
    entries = []
    for u, s in entries_query:
        # Create a combined object that matches the expected StatsEntry interface
        entry_data = {
            "division": u.division_name,
            "district": u.district_name,
            "upazila": u.name,
            "total": s.total if s else 0,
            "valid": s.valid if s else 0,
            "invalid": s.invalid if s else 0,
            "quota": u.quota or 0,
            "filename": s.filename if s else "",
            "version": s.version if s else 0,
            "created_at": (s.created_at if s else datetime.utcnow()),
            "updated_at": (s.updated_at if s else datetime.utcnow()),
            "pdf_url": s.pdf_url if s else "",
            "pdf_invalid_url": s.pdf_invalid_url if s else "",
            "excel_url": s.excel_url if s else "",
            "excel_valid_url": s.excel_valid_url if s else "",
            "excel_invalid_url": s.excel_invalid_url if s else "",
        }
        entries.append(entry_data)

    # 2. Cheap ETag based on latest updated_at + row count
    latest_ts = max((e['updated_at'] for e in entries), default=datetime.utcnow())
    etag_raw  = f"{len(entries)}:{latest_ts.isoformat()}"
    etag      = '"' + hashlib.md5(etag_raw.encode()).hexdigest() + '"'

    if request.headers.get("If-None-Match") == etag:
        from fastapi.responses import Response as FResponse
        return FResponse(status_code=304, headers={"ETag": etag})

    # 3. Master counts (derive from official tables)
    div_master  = db.query(Division.name, func.count(District.id)).join(District, Division.name == District.division_name).group_by(Division.name).all()
    div_counts  = {row[0]: row[1] for row in div_master}
    dist_master = db.query(District.name,  func.count(Upazila.id)).join(Upazila, District.name == Upazila.district_name).group_by(District.name).all()
    dist_counts = {row[0]: row[1] for row in dist_master}

    # 4. Grand total (Sum up the SummaryStats)
    grand = db.query(
        func.coalesce(func.sum(SummaryStats.total),   0).label("total"),
        func.coalesce(func.sum(SummaryStats.valid),    0).label("valid"),
        func.coalesce(func.sum(SummaryStats.invalid),  0).label("invalid"),
    ).first()
    grand_total = {"total": grand.total, "valid": grand.valid, "invalid": grand.invalid}

    from fastapi.responses import JSONResponse
    
    data = {
        "entries": [
            {
                **e,
                "created_at": e["created_at"].isoformat() + "Z",
                "updated_at": e["updated_at"].isoformat() + "Z",
            }
            for e in entries
        ],
        "grand_total": grand_total,
        "master_counts": {"divisions": div_counts, "districts": dist_counts},
        "last_modified": latest_ts.isoformat() + "Z",
    }
    return JSONResponse(
        content=data,
        headers={"ETag": etag, "Cache-Control": "private, no-store"},
    )

@app.delete("/statistics/{division}/{district}/{upazila}", dependencies=[Depends(PermissionChecker("view_admin"))])
async def delete_upazila_data(
    division: str,
    district: str,
    upazila: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Admin-only: Completely wipe all data for a specific upazila."""
    # 1. Delete ValidRecords
    db.query(ValidRecord).filter(
        ValidRecord.division == division,
        ValidRecord.district == district,
        ValidRecord.upazila == upazila
    ).delete(synchronize_session=False)

    # 2. Delete SummaryStats
    db.query(SummaryStats).filter(
        SummaryStats.division == division,
        SummaryStats.district == district,
        SummaryStats.upazila == upazila
    ).delete(synchronize_session=False)

    # 2.5 Delete InvalidRecords
    db.query(InvalidRecord).filter(
        InvalidRecord.division == division,
        InvalidRecord.district == district,
        InvalidRecord.upazila == upazila
    ).delete(synchronize_session=False)

    # 3. Delete UploadBatches
    db.query(UploadBatch).filter(
        UploadBatch.division == division,
        UploadBatch.district == district,
        UploadBatch.upazila == upazila
    ).delete(synchronize_session=False)

    db.commit()

    log_audit(
        db, 
        current_user, 
        "DELETE", 
        "upazila_full_wipe", 
        f"{division}/{district}/{upazila}", 
        old_data={"details": "Fully wiped records, stats, and batches for upazila"}
    )
    
    return {"status": "success", "message": f"All data for {upazila} has been deleted."}

@app.get("/search", dependencies=[Depends(PermissionChecker("view_stats"))])
async def search_records(
    query: str, 
    type: str = "nid", 
    page: int = 1,
    limit: int = 50,
    regex: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Search for valid records by NID, DOB, or Name with pagination and regex support."""
    query = query.strip()
    if not query:
        return {"results": [], "total": 0, "page": page, "limit": limit}
    
    limit = min(limit, 200)
    offset = (page - 1) * limit
    
    base_query = db.query(ValidRecord)
    
    if type == "dob":
        data_query = base_query.filter(ValidRecord.dob == query)
    elif type == "name":
        data_query = base_query.filter(ValidRecord.name.ilike(f"%{query}%"))
    else:  # default: nid
        if regex:
            # PostgreSQL case-insensitive regex operator
            data_query = base_query.filter(ValidRecord.nid.op("~*")(query))
        else:
            data_query = base_query.filter(ValidRecord.nid.contains(query))
            
    total = data_query.count()
    results = data_query.offset(offset).limit(limit).all()
    
    return {
        "results": results,
        "total": total,
        "page": page,
        "limit": limit
    }

@app.get("/nid/{nid}", dependencies=[Depends(PermissionChecker("view_stats"))])
async def check_nid(
    nid: str, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Public API: Check if an NID exists in the database. Returns the full record if found."""
    record = db.query(ValidRecord).filter(ValidRecord.nid == nid.strip()).first()
    if not record:
        raise HTTPException(status_code=404, detail=f"NID {nid} not found in database")
    return {
        "found": True,
        "nid": record.nid,
        "dob": record.dob,
        "name": record.name,
        "division": record.division,
        "district": record.district,
        "upazila": record.upazila,
        "source_file": record.source_file,
        "data": record.data,
        "created_at": record.created_at.isoformat() + "Z" if record.created_at else None,
        "updated_at": record.updated_at.isoformat() + "Z" if record.updated_at else None,
    }

@app.delete("/record/{record_id}", dependencies=[Depends(PermissionChecker("manage_users"))]) # Assuming admin role for deleting records
async def delete_record(
    record_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # Changed from admin: User = Depends(require_role(["admin"]))
):
    """Delete a single ValidRecord by ID. Decrements the related SummaryStats."""
    record = db.query(ValidRecord).filter(ValidRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    
    # Decrement SummaryStats for this record's location
    summary = db.query(SummaryStats).filter(
        SummaryStats.district == record.district,
        SummaryStats.upazila == record.upazila
    ).first()
    
    if summary:
        summary.valid = max(0, summary.valid - 1)
        summary.total = max(0, summary.total - 1)
    
    db.delete(record)
    db.commit()
    
    log_audit(db, current_user, "DELETE", "valid_records", record_id, old_data={"nid": record.nid})
    
    return {"deleted": True, "id": record_id, "nid": record.nid}

class ManualStatsUpdate(BaseModel):
    old_district: str
    old_upazila: str
    new_district: str
    new_upazila: str
    total: int
    valid: int
    invalid: int

@app.put("/statistics/update", dependencies=[Depends(PermissionChecker("manage_geo"))]) # Assuming admin role for updating stats
async def update_statistics(
    update: ManualStatsUpdate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user) # Changed from admin: User = Depends(require_role(["admin"]))
):
    """Manually update the statistics and location for a specific entry."""
    summary = db.query(SummaryStats).filter(
        SummaryStats.district == update.old_district,
        SummaryStats.upazila == update.old_upazila
    ).first()

    if not summary:
        raise HTTPException(status_code=404, detail="Statistic entry not found")

    summary.district = update.new_district
    summary.upazila = update.new_upazila
    summary.division = get_division_for_district(update.new_district)
    summary.total = update.total
    summary.valid = update.valid
    summary.invalid = update.invalid
    summary.version += 1
    
    db.commit()
    db.refresh(summary)
    
    log_audit(db, current_user, "UPDATE", "summary_stats", summary.id, new_data={"action": "summary_update"})
    
    return summary


@app.get("/admin/audit-logs", dependencies=[Depends(PermissionChecker("view_admin"))])
async def get_audit_logs(limit: int = 100, db: Session = Depends(get_db)):
    return db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit).all()

@app.get("/admin/api-usage", dependencies=[Depends(PermissionChecker("view_admin"))])
async def get_api_usage(limit: int = 100, db: Session = Depends(get_db)):
    return db.query(ApiUsageLog).order_by(ApiUsageLog.created_at.desc()).limit(limit).all()


@app.get("/geo-info", dependencies=[Depends(PermissionChecker("view_geo"))])
async def get_geo_info(db: Session = Depends(get_db)):
    """Return the hierarchy of divisions, districts, and upazilas."""
    from .bd_geo import _division_lookup, _district_lookup, get_dynamic_upazilas
    
    divisions = list(_division_lookup.keys())
    # The keys in _division_lookup are normalized (lowercase).
    # Since the UI expects proper names, let's just get them from the values.
    divisions = sorted(list(set(_division_lookup.values())))
    
    districts = {}
    for norm_name, record in _district_lookup.items():
        div_name = record["division_name"]
        dist_name = record["name"]
        if div_name not in districts:
            districts[div_name] = []
        if dist_name not in districts[div_name]:
            districts[div_name].append(dist_name)
            
    for div in districts:
        districts[div].sort()
    
    upazilas = get_dynamic_upazilas(db)
        
    return {
        "divisions": divisions,
        "districts": districts,
        "upazilas": upazilas
    }

class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

@app.post("/auth/change-password")
async def change_password(req: ChangePasswordRequest, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not verify_password(req.old_password, current_user.hashed_password):
        raise HTTPException(status_code=400, detail="Incorrect old password")
    
    current_user.hashed_password = get_password_hash(req.new_password)
    db.commit()
    
    log_audit(db, current_user, "UPDATE", "users", current_user.id, new_data={"action": "password_change"})
    
    return {"message": "Password updated successfully"}

@app.get("/auth/me")
async def read_users_me(current_user: User = Depends(get_current_user)):
    return current_user

@app.get("/guess-location")
async def guess_location(filename: str):
    """Guess the location from the filename."""
    from .bd_geo import fuzzy_match_location
    return fuzzy_match_location(filename)

# --- NID Verification & Sync ---

@app.get("/ibas/nid-verify")
@limiter.limit(get_db_rate_limit)
async def ibas_verify_nid(
    request: Request,
    id: str,
    db: Session = Depends(get_db),
    api_user: User = Depends(get_api_key)
):
    """Secure, rate-limited endpoint for IBAS to verify NID."""
    if not id or not id.strip():
        raise HTTPException(status_code=400, detail="Missing NID")
    
    # Simple check, we don't return PII for security, just boolean
    exists = db.query(ValidRecord).filter(ValidRecord.nid == id.strip()).first() is not None
    return {"found": exists}

@app.get("/sync/export")
async def sync_export(
    since: datetime = None,
    db: Session = Depends(get_db),
    api_user: User = Depends(get_api_key)
):
    """Export ValidRecords modified after the provided timestamp."""
    query = db.query(ValidRecord)
    if since:
        query = query.filter(ValidRecord.updated_at > since)
    
    # We might want to limit the payload size in production, but for now we export all requested
    records = query.all()
    out = []
    for r in records:
        out.append({
            "nid": r.nid,
            "dob": r.dob,
            "name": r.name,
            "division": r.division,
            "district": r.district,
            "upazila": r.upazila,
            "source_file": r.source_file,
            "upload_batch": r.upload_batch,
            "data": r.data,
            "created_at": r.created_at.isoformat() + "Z" if r.created_at else None,
            "updated_at": r.updated_at.isoformat() + "Z" if r.updated_at else None,
        })
    return {"records": out}

@app.post("/sync/import")
async def sync_import(
    payload: dict,
    db: Session = Depends(get_db),
    api_user: User = Depends(get_api_key)
):
    """Import ValidRecords via bulk upsert."""
    records = payload.get("records", [])
    if not records:
        return {"imported": 0}

    insert_data = []
    for r in records:
        created_at = datetime.fromisoformat(r["created_at"].rstrip("Z")) if r.get("created_at") else datetime.utcnow()
        updated_at = datetime.fromisoformat(r["updated_at"].rstrip("Z")) if r.get("updated_at") else datetime.utcnow()
        insert_data.append({
            "nid": r["nid"],
            "dob": r.get("dob", ""),
            "name": r.get("name", ""),
            "division": r.get("division", ""),
            "district": r.get("district", ""),
            "upazila": r.get("upazila", ""),
            "source_file": r.get("source_file", ""),
            "upload_batch": r.get("upload_batch", 1),
            "data": r.get("data", {}),
            "created_at": created_at,
            "updated_at": updated_at
        })

    # Chunked upsert
    for i in range(0, len(insert_data), 1000):
        chunk = insert_data[i:i+1000]
        stmt = insert(ValidRecord).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=['nid'],
            set_={
                'dob': stmt.excluded.dob,
                'name': stmt.excluded.name,
                'division': stmt.excluded.division,
                'district': stmt.excluded.district,
                'upazila': stmt.excluded.upazila,
                'source_file': stmt.excluded.source_file,
                'upload_batch': stmt.excluded.upload_batch,
                'data': stmt.excluded.data,
                'updated_at': stmt.excluded.updated_at
            }
        )
        db.execute(stmt)
    db.commit()
    return {"imported": len(records)}

import httpx

@app.post("/admin/instances/{id}/trigger-sync", dependencies=[Depends(PermissionChecker("view_admin"))])
async def trigger_remote_sync(
    id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from .models import RemoteInstance
    instance = db.query(RemoteInstance).filter(RemoteInstance.id == id).first()
    if not instance:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    since_param = f"?since={instance.last_synced_at.isoformat()}" if instance.last_synced_at else ""
    url = f"{instance.url.rstrip('/')}/sync/export{since_param}"
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.get(url, headers={"X-API-Key": instance.api_key})
            resp.raise_for_status()
            data = resp.json()
            records = data.get("records", [])
            
            # Use our own import function to dry run the code
            if records:
                await sync_import({"records": records}, db=db, api_user=current_user)
            
            instance.last_synced_at = datetime.utcnow()
            db.commit()
            return {"synced_count": len(records), "status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")

if __name__ == "__main__":
    uvicorn.run("app.main:app", host="0.0.0.0", port=8000, reload=True)
