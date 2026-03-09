import pandas as pd
import json

file_path = "c:/FFP-DataValidation/Cumilla_Brammanpara - DC Food Comilla (1).xlsx"

# Let's see rows 0-10 using openpyxl just to see what they look like
from openpyxl import load_workbook
wb = load_workbook(file_path, data_only=True)
ws = wb.active

print("--- RAW ROWS ---")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i < 10:
        print(f"Row {i+1}: {row}")

# Now let's try reading it with pandas with different headers
print("\n--- PANDAS HEADERS ---")
for h in range(5):
    try:
        df = pd.read_excel(file_path, header=h, dtype=str)
        print(f"header={h}: columns={df.columns.tolist()}")
    except Exception as e:
        print(f"header={h} failed: {e}")
